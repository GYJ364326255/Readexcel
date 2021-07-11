#-*- codeing = utf-8 -*-   
#@Time : 09/07/2021 10:48
#@Author ： 高永佳
#@File : Exceltabelle.py
#@Software: PyCharm

from openpyxl import Workbook
from openpyxl import load_workbook

#read excelfile, return the data of excel file in a way of list
def read_excelfile(file,startcolumn,endcolumn,startrow,endrow):
    file = file.active
    wholedata = []
    for i in range(startcolumn,endcolumn+1):
        subdata = []
        for j in range(startrow,endrow+1):
            subdata.append((file.cell(row=j,column=i)).value)
        wholedata.append(subdata)
    print(len(wholedata))
    return wholedata

#delete column, with the special condition
def delete_column(lst,zeronumb): #zeronum is the number of zero cell in the excel table in each column
    #print(len(lst))
    modlst = []
    for x in range(len(lst)): #not a array
        num = lst[x].count(None) #not such list.find(only return index)/ None has meaning in python[None]
        print(num)
        if num < zeronumb:
            # lst.pop(x) #pop the index with x; pop之后就导致程序的列数不等了
            modlst.append(lst[x])
    return modlst #here you have to return value otherwise did not success

#write the data into excel
def write_excel(Datalist,sheettitle='Sheet'):
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheettitle
    for i in range(len(Datalist[0])): #column number: the number of list element in the datalist
        Sublist = []
        for j in range(len(Datalist)): #row number: the number of list element in firt sublist
            Sublist.append(Datalist[j][i])
        print('line',i)
        print(Sublist)
        sheet.append(Sublist)
    Myexcelfile = r'modifleddata.xlsx'
    wb.save(filename=Myexcelfile)


ExcelFile = load_workbook('excelfilemodified.xlsx')
wholedata = read_excelfile(ExcelFile,1,22,1,21)
filterdata = delete_column(wholedata,2)
write_excel(filterdata)

print('#'*50 + 'this is the new data')
print(filterdata)


